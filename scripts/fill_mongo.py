from pymongo import MongoClient
from openpyxl import load_workbook
import sys

files = ['../Excels/Criminal_Offenses_Local_State_Police.xlsx',
			'../Excels/Criminal_Offenses_Noncampus.xlsx',
			'../Excels/Criminal_Offenses_On_campus.xlsx',
			'../Excels/Criminal_Offenses_Public_Property.xlsx']

client = MongoClient()
db = client.campus_crime

year = sys.argv[1];
collection_name = "totals" + year;
campus_stats = db[collection_name]

year = int(year)

for file in files:
	wb = load_workbook(file)
	ws = wb.active

	#total
	for row in ws.iter_rows() :
		row_year = row[0].value
		if type(row_year) is str :
			continue

		#get total (forcible and non forcible offenses)
		total = row[6].value + row[9].value
		
		if(row_year == year and total > 0):
			#print(row[2].value)
			#campus_id = None

			#for 2014 only
			#rapes = int(row[7].value) + int(row[9].value)

			#only needed for first run (2014)
			'''
			#add to campuses document
			cursor = db.campuses.find({"school_name" : row[2].value, "campus_name" : row[4].value})
			if cursor.count() == 0 :
				result = db.campuses.insert_one(
					{
						"school_id" : row[1].value,
						"school_name" : row[2].value,
						"campus_name" : row[4].value,
						"lat" : 0,
						"long" : 0,
					}
					)
				campus_id = result.inserted_id;
			elif cursor.count() == 1 :
				campus_id = cursor[0]['_id']
			'''

			#add/update totals collection
			cursor = campus_stats.find({"school_name" : row[2].value, "campus_name" : row[4].value})
			if cursor.count() == 1 :
				#add to running total
				result = campus_stats.update(
						{"school_name" : row[2].value, "campus_name" : row[4].value},
						{ "$inc" : { "total_offenses" : total} }
					)
				continue
			elif cursor.count() == 0:
				#insert
				result = campus_stats.insert_one(
					{
						"school_id" : row[1].value,
						"school_name" : row[2].value,
						"campus_name" : row[4].value,
						"size" : row[5].value,
						"total_offenses" : total,
					}
					)
